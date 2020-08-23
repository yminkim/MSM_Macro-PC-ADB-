#엑셀 컬럼을 긁어와서 리스트에 저장시키는 코드

from openpyxl import load_workbook
load_wb = load_workbook("C:/Users/user/Desktop/reserve.xlsx", data_only=True)
load_ws = load_wb['Sheet1']
print(load_ws['A1'].value)

reserve_list = []
a = range(0,15)
for i in a:
    i +=1
    reserve_list.append(load_ws.cell(i,1).value)

for j in a:
    print(reserve_list[j]) 
    j +=1
