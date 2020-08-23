#컬럼값 긁어와서 메모장에 붙여넣기

#pip install pyautogui  


import pyautogui 
from openpyxl import load_workbook

load_wb = load_workbook("C:/Users/user/Desktop/reserve.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

reserve_list = []
a = range(0,15)

for i in a:
    i +=1
    reserve_list.append(load_ws.cell(i,2).value)

pyautogui.moveTo(1133, 479, 0.1)
pyautogui.click(button='left')

for j in a:
    j+=1 
    pyautogui.write(reserve_list[j])
    pyautogui.press('enter')
