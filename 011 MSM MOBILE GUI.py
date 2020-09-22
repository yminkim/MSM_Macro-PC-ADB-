pip install -U pure-python-adb

from ppadb.client import Client as AdbClient
from bs4 import BeautifulSoup
import re
from tkinter import*
import pyautogui
import pyperclip
from tkinter import filedialog
from openpyxl import load_workbook
import os
import tkinter.messagebox
import time
import threading

    
###############################################
win = Tk()
win.geometry("460x460")
win.title("메엠 템세팅 폰크로")
win.option_add("*Font","맑은고딕 10")
###############################################


global filepath
global file_Sheet


client = AdbClient(host="127.0.0.1", port=5037)
device = client.devices()[0]



def infoBox():
    asdfoijn = '''작성중,
    일시정지는 케이블 뽑으면 됩니다'''
    tkinter.messagebox.showinfo("설명서", asdfoijn)
###############################################
def note8():
    type_here_xlct.insert(0,"1800")
    type_here_ylct.insert(0,"342")
    backSpace_xlct.insert(0,"2422")
    backSpace_ylct.insert(0,"1219")
    execute_xlct.insert(0,"2300")
    execute_ylct.insert(0,"1331")
    
###############################################
def clear():
    type_here_xlct.delete(0,1000)
    type_here_ylct.delete(0,1000)
    backSpace_xlct.delete(0,1000)
    backSpace_ylct.delete(0,1000)
    execute_xlct.delete(0,1000)
    execute_ylct.delete(0,1000)
    #filepath  = None
    #fileLabel.config(text = "↑ 파일첨부 해주세요")

###############################################
def update_device():
    global lbl
    global text
    while True:
        text.set(client.devices())
        time.sleep(1)
        
###############################################    
def openFile():
    global filepath
    filepath = filedialog.askopenfilename()
    fileLabel.config(text = os.path.basename(filepath))
    print(filepath)
###############################################
###############################################
def runas():
    TYPE_HERE_X = int(type_here_xlct.get())
    TYPE_HERE_Y = int(type_here_ylct.get())

    BACK_SPACE_X = int(backSpace_xlct.get())
    BACK_SPACE_Y = int(backSpace_ylct.get())

    #COMPLETE_X = int(complete_xlct.get())
    #COMPLETE_Y = int(complete_ylct.get())

    EXECUTE_X = int(execute_xlct.get())
    EXECUTE_Y = int(execute_ylct.get())
    
    global filepath
    fise_Sheet = sheetEnt.get()
    
    load_wb = load_workbook(filepath, data_only=True)
    load_ws = load_wb[fise_Sheet]
    
    index_list = []
    cul = int(culumm_Ent.get()) -1
    row = int(row_Ent.get())
    
    cul2 = int(culumm_Ent2.get())
    cul_range = (cul2 - cul)+1
    
    a = range(0,cul_range)
    
    for i in a:
        index_list.append(load_ws.cell(cul,row).value)
        cul +=1
       
    for j in a:
        j+=1 
        
        device.shell(f'input tap {TYPE_HERE_X} {TYPE_HERE_Y}');
        time.sleep(0.07)
        device.shell(f'input swipe {BACK_SPACE_X} {BACK_SPACE_Y} {BACK_SPACE_X} {BACK_SPACE_Y} 1000')
        time.sleep(0.07)
        device.shell(f'input text {index_list[j]}')  
        time.sleep(0.07)
        #device.shell(f'input tap {COMPLETE_X} {COMPLETE_Y}'); 
        device.shell('input keyevent 66');
        time.sleep(0.07)
        device.shell(f'input tap {EXECUTE_X} {EXECUTE_Y}');
        time.sleep(0.07)
        
        
    
###############################################
# 세팅 라벨
mobileSET_Label = Label(win)
mobileSET_Label.config(text = "[Setting]")
mobileSET_Label.config(font=("맑은고딕", 15))
mobileSET_Label.place(x=20, y=25)
mobileSET_Label.place()

#type here좌표 라벨
type_here_label = Label(win)
type_here_label.config(text = "▶ Type Here...")
type_here_label.place(x=25, y=55)
type_here_label.place()

#type here좌표 라벨
type_here_xlabel = Label(win)
type_here_xlabel.config(text = "X:")
type_here_xlabel.place(x=125, y=55)
type_here_xlabel.place()

#nox x 입력창
type_here_xlct = Entry(win)
type_here_xlct.place(x=145, y=55, width=40, height=18)
type_here_xlct.place()

#nox  y좌표 라벨
type_here_ylabel = Label(win)
type_here_ylabel.config(text = "Y:")
type_here_ylabel.place(x=190, y=55)
type_here_ylabel.place()

#nox y 입력창
type_here_ylct = Entry(win)
type_here_ylct.place(x=210, y=55, width=40, height=18)
type_here_ylct.place()

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@###


#Back space 좌표 라벨
backSpace_label = Label(win)
backSpace_label.config(text = "▶ Backspace")
backSpace_label.place(x=25, y=75)
backSpace_label.place()

#type here좌표 라벨
backSpace_xlabel = Label(win)
backSpace_xlabel.config(text = "X:")
backSpace_xlabel.place(x=125, y=75)
backSpace_xlabel.place()

#nox x 입력창
backSpace_xlct = Entry(win)
backSpace_xlct.place(x=145, y=75, width=40, height=18)
backSpace_xlct.place()

#nox  y좌표 라벨
backSpace_ylabel = Label(win)
backSpace_ylabel.config(text = "Y:")
backSpace_ylabel.place(x=190, y=75)
backSpace_ylabel.place()

#nox y 입력창
backSpace_ylct = Entry(win)
backSpace_ylct.place(x=210, y=75, width=40, height=18)
backSpace_ylct.place()

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@###
#Back space 좌표 라벨
#complete_label = Label(win)
#complete_label.config(text = "▶ 완료")
#complete_label.place(x=25, y=95)
#complete_label.place()

#type here좌표 라벨
#complete_xlabel = Label(win)
#complete_xlabel.config(text = "X:")
#complete_xlabel.place(x=125, y=95)
#complete_xlabel.place()

#nox x 입력창
#complete_xlct = Entry(win)
#complete_xlct.place(x=145, y=95, width=40, height=18)
#complete_xlct.place()

#nox  y좌표 라벨
#complete_ylabel = Label(win)
#complete_ylabel.config(text = "Y:")
#complete_ylabel.place(x=190, y=95)
#complete_ylabel.place()

#nox y 입력창
#complete_ylct = Entry(win)
#complete_ylct.place(x=210, y=95, width=40, height=18)
#complete_ylct.place()



#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@###
#Execute좌표 라벨
execute_label = Label(win)
execute_label.config(text = "▶ Execute")
execute_label.place(x=25, y=95)
execute_label.place()

#type here좌표 라벨
execute_xlabel = Label(win)
execute_xlabel.config(text = "X:")
execute_xlabel.place(x=125, y=95)
execute_xlabel.place()

#nox x 입력창
execute_xlct = Entry(win)
execute_xlct.place(x=145, y=95, width=40, height=18)
execute_xlct.place()

#nox  y좌표 라벨
execute_ylabel = Label(win)
execute_ylabel.config(text = "Y:")
execute_ylabel.place(x=190, y=95)
execute_ylabel.place()

#nox y 입력창
execute_ylct = Entry(win)
execute_ylct.place(x=210, y=95, width=40, height=18)
execute_ylct.place()

###############################################
#파일 찾는 버튼        
fileLoad_btn = Button(text = "파일찾기", command = openFile)
fileLoad_btn.place(x=25, y=160)
fileLoad_btn.place()

#불러온 파일 이름 라벨
fileLabel2 = Label(win)
fileLabel2.config(text = "파일:")
fileLabel2.place(x=25, y=185)
fileLabel2.place()
#불러온 파일 이름 라벨2
fileLabel = Label(win)
fileLabel.config(text = "↑ 파일첨부 해주세요")
fileLabel.place(x=60, y=185)
fileLabel.place()

###############################################
#시트 라벨
sheetLabel = Label(win)
sheetLabel.config(text = "시트명:")
sheetLabel.place(x=97, y=163)
sheetLabel.place()

#시트 입력창
sheetEnt = Entry(win)
sheetEnt.insert(0,"Sheet1")
sheetEnt.place(x=148, y=165, width=90, height=15)
sheetEnt.place()
###############################################.
#범위 지정 라벨
range_label = Label(win)
range_label.config(text = "범위 지정(※행은 최소 2이상 입력)")
range_label.place(x=24, y=215)
range_label.place()

#행 라벨
culumm_label = Label(win)
culumm_label.config(text = " 행  :")
culumm_label.place(x=26, y=235)
culumm_label.place()

#행 입력창
culumm_Ent = Entry(win)
culumm_Ent.place(x=63, y=235, width=35, height=18)
culumm_Ent.place()

#행2 라벨
culumm_label2 = Label(win)
culumm_label2.config(text = "~")
culumm_label2.place(x=101, y=235)
culumm_label2.place()

#행2 입력창
culumm_Ent2 = Entry(win)
culumm_Ent2.place(x=120, y=235, width=35, height=18)
culumm_Ent2.place()

#열 라벨
row_label = Label(win)
row_label.config(text = " 열  :")
row_label.place(x=26, y=255)
row_label.place()

#열 입력창
row_Ent = Entry(win)
row_Ent.place(x=63, y=255, width=35, height=18)
row_Ent.place()

###############################################.
device_info = Label(win)
device_info.config(text = "Device:")
device_info.place(x=5, y=0)
device_info.place()
###############################################.
text = StringVar()
lbl = Label(win, textvariable=text)
lbl.place(x=55, y=0)
lbl.place()

t = threading.Thread(target=update_device)
t.daemon = True
t.start()
###############################################.
btn_info = Button(win)
btn_info.config(text="도움")
#btn_info.config(width = 2, height=1)
btn_info.place(x=30, y=420)
btn_info.config(command = infoBox)
btn_info.place()

#실행
runas_btn = Button(win)
runas_btn.config(text="실행")
runas_btn.config(width = 6, height=1)
runas_btn.place(x=170, y=420)
runas_btn.config(command = runas)
runas_btn.place()

#초기화
clear_btn = Button(win)
clear_btn.config(text="초기화")
clear_btn.config(width = 6, height=1)
clear_btn.place(x=236, y=420)
clear_btn.config(command = clear)
clear_btn.place()

#메모장 라벨
mouse_location = Label(win)
mouse_location.config(text = "메모장")
mouse_location.place(x=24, y=286)
mouse_location.place()
###############################################################
mobilelist_Label = Label(win)
mobilelist_Label.config(text = "[MOBILE LIST]")
mobilelist_Label.place(x=295, y=35)
mobilelist_Label.place()


device_note8_btn = Button(win)
device_note8_btn.config(text="Note8(2960 x 1440)")
device_note8_btn.config(width = 15, height=1)
device_note8_btn.place(x=300, y=55)
device_note8_btn.config(command = note8)
device_note8_btn.place()

################################################################


textBox = Text(win)
textBox.place(x=26, y=305, width=269, height=100)
textBox.place()


win.mainloop()
