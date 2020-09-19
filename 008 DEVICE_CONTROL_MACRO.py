pip install -U pure-python-adb
from ppadb.client import Client as AdbClient
from bs4 import BeautifulSoup
import re
from tkinter import*
import pyautogui
import pyperclip
from tkinter import filedialog
from openpyxl import load_workbook
import os
import tkinter.messagebox
import time
import threading

client = AdbClient(host="127.0.0.1", port=5037)
device = client.devices()[0]

load_wb = load_workbook('C:/Users/user/Desktop/reserve.xlsx', data_only=True)
load_ws = load_wb['Sheet1']
reserve_list = []

a = range(0,10)
b = range(0,5)
for i in a:
    i +=1
    reserve_list.append(load_ws.cell(i,2).value)

for j in a:
    j+=1
    device.shell('input tap 1800 350');
    time.sleep(0.07)
    device.shell('input swipe 2440 1221 2440 1221 1000')
    #for k in b:
    #    device.shell('input keyevent 67')

    time.sleep(0.07)
    device.shell(f'input text {reserve_list[j]}')  
    time.sleep(0.07)
    device.shell('input tap 2384 1350');  
    time.sleep(0.07)
    device.shell('input tap 2269 1341'); 
    time.sleep(0.07)
