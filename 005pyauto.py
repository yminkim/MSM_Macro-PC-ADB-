# 녹스에 입력하고 지웠다가 하는 그런 코드임


import pyperclip
import pyautogui 
from openpyxl import load_workbook

load_wb = load_workbook("C:/Users/user/Desktop/(파일이름).xlsx", data_only=True)
load_ws = load_wb['Sheet1']

index_list = []
a = range(0,50)

for i in a:
    i +=1
    index_list.append(load_ws.cell(i,1).value)

pyautogui.moveTo(1073, 61, 0.1) #녹스 위치
pyautogui.click(button='left')

for j in a:
    j+=1 
    pyautogui.press('h') #input칸 녹스의 단축키
    for k in range(0,5):
        pyautogui.press('backspace') // 입력한 인덱스를 지워줌
    pyperclip.copy(index_list[j])
    pyautogui.hotkey("ctrl", "v")
    #pyautogui.write(reserve_list[j])
    pyautogui.press('enter')
    pyautogui.press('j')
