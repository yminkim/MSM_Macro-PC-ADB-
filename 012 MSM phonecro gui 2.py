from ppadb.client import Client as AdbClient
from bs4 import BeautifulSoup
import re
from tkinter import*
import pyautogui
import pyperclip
from tkinter import filedialog
from openpyxl import load_workbook
import os
import tkinter.messagebox
import time
import threading


win = Tk()
win.geometry("460x460")
win.title("메엠 템세팅 폰크로")
win.option_add("*Font","맑은고딕 10")
global Device_filepath
global Device_file_Sheet


def Device_openFile():
    global Device_filepath
    Device_filepath = filedialog.askopenfilename()
    Device_fileLabel.config(text = os.path.basename(Device_filepath))
    print(Device_filepath)


def device_location():
    global Device_filepath
    Device_fise_Sheet = Device_sheetEnt.get()

    Device_load_wb = load_workbook(Device_filepath, data_only=True)
    Device_load_ws = Device_load_wb[Device_fise_Sheet]
    phone_device_row = int(device_num_Ent.get())
    #pppp = 3
    #qqqqq = 1
    #device_location_range = range(0,6)
    #for i in device_location_range :
        #location_[qqqqq].insert(load_ws.cell(device_location_range,phone_device_row).value)
        #pppp+=1
        #qqqqq +=1
    
    type_here_xlct.delete(0,1000)
    type_here_ylct.delete(0,1000)
    backSpace_xlct.delete(0,1000)
    backSpace_ylct.delete(0,1000)
    execute_xlct.delete(0,1000)
    execute_ylct.delete(0,1000)
    
    lct_num1 = Device_load_ws.cell(phone_device_row,3).value
    lct_num2 = Device_load_ws.cell(phone_device_row,4).value
    lct_num3 = Device_load_ws.cell(phone_device_row,5).value
    lct_num4 = Device_load_ws.cell(phone_device_row,6).value
    lct_num5 = Device_load_ws.cell(phone_device_row,7).value
    lct_num6 = Device_load_ws.cell(phone_device_row,8).value
    
    type_here_xlct.insert(0, lct_num1)
    type_here_ylct.insert(0, lct_num2)
    backSpace_xlct.insert(0, lct_num3)
    backSpace_ylct.insert(0, lct_num4)
    execute_xlct.insert(0, lct_num5)
    execute_ylct.insert(0, lct_num6)
    


#파일 찾는 버튼        
Device_fileLoad_btn = Button(text = "파일찾기", command = Device_openFile)
Device_fileLoad_btn.place(x=25, y=160)
Device_fileLoad_btn.place()

#불러온 파일 이름 라벨
Device_fileLabel2 = Label(win)
Device_fileLabel2.config(text = "파일:")
Device_fileLabel2.place(x=25, y=185)
Device_fileLabel2.place()
#불러온 파일 이름 라벨2
Device_fileLabel = Label(win)
Device_fileLabel.config(text = "↑ 파일첨부 해주세요")
Device_fileLabel.place(x=60, y=185)
Device_fileLabel.place()

#시트 라벨
Device_sheetLabel = Label(win)
Device_sheetLabel.config(text = "시트명:")
Device_sheetLabel.place(x=97, y=163)
Device_sheetLabel.place()

#시트 입력창
Device_sheetEnt = Entry(win)
Device_sheetEnt.insert(0,"Sheet1")
Device_sheetEnt.place(x=148, y=165, width=90, height=15)
Device_sheetEnt.place()



#행 라벨
sheetLabel = Label(win)
sheetLabel.config(text = "디바이스행:")
sheetLabel.place(x=70, y=205)
sheetLabel.place()
#행 번호 입력칸d
device_num_Ent = Entry(win)
device_num_Ent.place(x=148, y=205, width=90, height=15)
device_num_Ent.place()

#실행
device_lct_btn = Button(win)
device_lct_btn.config(text="좌표값")
device_lct_btn.config(width = 6, height=1)
device_lct_btn.place(x=148, y=225)
device_lct_btn.config(command = device_location)
device_lct_btn.place()

/*
#좌표 1
location_1 = Entry(win)
location_1.place(x=148, y=245, width=90, height=15)
location_1.place()

#좌표 2
location_2 = Entry(win)
location_2.place(x=148, y=265, width=90, height=15)
location_2.place()

#좌표 3
location_3 = Entry(win)
location_3.place(x=148, y=285, width=90, height=15)
location_3.place()

#좌표 4
location_4 = Entry(win)
location_4.place(x=148, y=305, width=90, height=15)
location_4.place()

#좌표 5
location_5 = Entry(win)
location_5.place(x=148, y=325, width=90, height=15)
location_5.place()

#좌표 6
location_6 = Entry(win)
location_6.place(x=148, y=345, width=90, height=15)
location_6.place()
*/




win.mainloop()
