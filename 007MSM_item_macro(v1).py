#runas에서 딜레이 추가 및 인덱스 길이에 맞춰 backspace 입력
# sleep 대체 함수 확인 


from tkinter import*
import pyautogui
import pyperclip
from tkinter import filedialog
from openpyxl import load_workbook
import os
import tkinter.messagebox
import time
import threading
###############################################
win = Tk()
win.geometry("320x460")
win.title("0000")
win.option_add("*Font","맑은고딕 10")
###############################################
global filepath
global file_Sheet
###############################################
def updateMousePos():
    global lbl
    global text
    while True:
        x, y = pyautogui.position()
        text.set(f"x: {x}  y: {y} ")
        time.sleep(0.05)
        
###############################################        
def infoBox():
    asdfoijn = '''[매크로 설명서]
 1. 녹스 좌표(x, y): 녹스 인게임 화면 좌표 입력
 2. 단축키 지정: 단축키로 지정해둔 문자 입력
 3. 파일찾기: 사용할 파일 가져오기
 4. 시트명: 사용할 시트의 이름 입력
 5. 범위지정: 
  행-1) 범위의 시작 행 입력 (단, 2이상으로 입력)
  행-2) 범위의 마직막 행 입력
  열-1) 범위 열 입력
    > 열은 a-1, b-2, c-3, d-4 순서대로 숫자로 입력
      (ex. a열 = 1, h열 = 8)

[★주의사항★]
 - 매크로 작동 중 별도의 클릭을 하시면 안 됩니다.
   
 - item 'Type Here' 박스는 공백상태로 실행합니다.
 
 - 파일 및 시트 볼륨이 클 경우 
   파일을 읽는 속도가 느리거나 실행이 불가합니다.
   가급적이면 새로운 파일에서 사용할 값만 
   추출한 시트 사용을 추천드립니다.

 - 유효하지 않은값(공란, 잘못된 index)이 포함된 경우 
   해당 부분에서 작동이 중지될 수 있습니다.'''
    tkinter.messagebox.showinfo("설명서", asdfoijn)
###############################################
def openFile():
    global filepath
    filepath = filedialog.askopenfilename()
    fileLabel.config(text = os.path.basename(filepath))
    print(filepath)
###############################################
def clear():
    nox_x.delete(0,1000)
    nox_y.delete(0,1000)
    sheetEnt.delete(0,1000)
    filepath  = None
    fileLabel.config(text = "↑ 파일첨부 해주세요")
    culumm_Ent.delete(0,1000)
    culumm_Ent2.delete(0,1000)
    row_Ent.delete(0,1000)
###############################################
def runas():
    global filepath
    fise_Sheet = sheetEnt.get()
    xLct = nox_x.get()
    yLct = nox_y.get()
    
    item_val = item_ent.get()
    execute_val = execute_ent.get()
    
    load_wb = load_workbook(filepath, data_only=True)
    load_ws = load_wb[fise_Sheet]
    
    index_list = []
    cul = int(culumm_Ent.get()) -1
    row = int(row_Ent.get())
    
    cul2 = int(culumm_Ent2.get())
    cul_range = (cul2 - cul)+1
    
    a = range(0,cul_range)
    
    for i in a:
        index_list.append(load_ws.cell(cul,row).value)
        cul +=1
        
    pyautogui.moveTo(int(xLct), int(yLct), 0.1) #위치
    pyautogui.click(button='left')
    
    pyautogui.press(item_val)    
    for j in a:
        j+=1 
        time.sleep(0.07)
        pyperclip.copy(index_list[j])
        index_length = len(str(index_list[j]))
        pyautogui.hotkey("ctrl", "v")
        time.sleep(0.07)
        pyautogui.press('enter')
        time.sleep(0.07)
        pyautogui.press(execute_val)
        time.sleep(0.07)
        pyautogui.press(item_val)
        time.sleep(0.07)
        for k in range(0,index_length):
            pyautogui.press('backspace')

    
###############################################
#파일 찾는 버튼        
fileLoad_btn = Button(text = "파일찾기", command = openFile)
fileLoad_btn.place(x=25, y=160)
fileLoad_btn.place()

#불러온 파일 이름 라벨
fileLabel2 = Label(win)
fileLabel2.config(text = "파일:")
fileLabel2.place(x=25, y=185)
fileLabel2.place()
#불러온 파일 이름 라벨2
fileLabel = Label(win)
fileLabel.config(text = "↑ 파일첨부 해주세요")
fileLabel.place(x=60, y=185)
fileLabel.place()

###############################################
#시트 라벨
sheetLabel = Label(win)
sheetLabel.config(text = "시트명:")
sheetLabel.place(x=97, y=163)
sheetLabel.place()

#시트 입력창
sheetEnt = Entry(win)
sheetEnt.insert(0,"Sheet1")
sheetEnt.place(x=148, y=165, width=90, height=15)
sheetEnt.place()
###############################################.
#범위 지정 라벨
range_label = Label(win)
range_label.config(text = "범위 지정(※행은 최소 2이상 입력)")
range_label.place(x=24, y=215)
range_label.place()

#행 라벨
culumm_label = Label(win)
culumm_label.config(text = " 행  :")
culumm_label.place(x=26, y=235)
culumm_label.place()

#행 입력창
culumm_Ent = Entry(win)
culumm_Ent.place(x=63, y=235, width=35, height=18)
culumm_Ent.place()

#행2 라벨
culumm_label2 = Label(win)
culumm_label2.config(text = "~")
culumm_label2.place(x=101, y=235)
culumm_label2.place()

#행2 입력창
culumm_Ent2 = Entry(win)
culumm_Ent2.place(x=120, y=235, width=35, height=18)
culumm_Ent2.place()

#열 라벨
row_label = Label(win)
row_label.config(text = " 열  :")
row_label.place(x=26, y=255)
row_label.place()

#열 입력창
row_Ent = Entry(win)
row_Ent.place(x=63, y=255, width=35, height=18)
row_Ent.place()

###############################################.
#녹스 세팅 라벨
NOXSET_Label = Label(win)
NOXSET_Label.config(text = "[Setting]")
NOXSET_Label.config(font=("맑은고딕", 15))
NOXSET_Label.place(x=20, y=25)
NOXSET_Label.place()

#nox 좌표 라벨
no_location = Label(win)
no_location.config(text = "녹스 좌표")
no_location.place(x=30, y=55)
no_location.place()

#nox x좌표 라벨
nox_location = Label(win)
nox_location.config(text = "X:")
nox_location.place(x=110, y=55)
nox_location.place()

#nox x 입력창
nox_x = Entry(win)
nox_x.place(x=130, y=55, width=40, height=18)
nox_x.place()

#nox  y좌표 라벨
noy_location = Label(win)
noy_location.config(text = "Y:")
noy_location.place(x=175, y=55)
noy_location.place()

#nox y 입력창
nox_y = Entry(win)
nox_y.place(x=195, y=55, width=40, height=18)
nox_y.place()
###############################################.
#아이템 단축키 라벨
dan_label = Label(win)
dan_label.config(text = "단축키 지정")
dan_label.place(x=30, y=80)
dan_label.place()

#아이템 단축키 라벨
item_dan = Label(win)
item_dan.config(text = "▶ item")
item_dan.place(x=36, y=100)
item_dan.place()

#아이템 단축키 입력
item_ent = Entry(win)
item_ent.place(x=130, y=100, width=25, height=18)
item_ent.place()

#확인버튼 단축키 라벨
item_dan = Label(win)
item_dan.config(text = "▶ Execute")
item_dan.place(x=36, y=120)
item_dan.place()

#확인버튼 단축키 입력
execute_ent = Entry(win)
execute_ent.place(x=130, y=120, width=25, height=18)
execute_ent.place()

###############################################.

#메모장 라벨
mouse_location = Label(win)
mouse_location.config(text = "메모장")
mouse_location.place(x=24, y=286)
mouse_location.place()


textBox = Text(win)
textBox.place(x=26, y=305, width=269, height=100)
textBox.place()

###############################################.
text = StringVar()
lbl = Label(win, textvariable=text)
lbl.place(x=195, y=10)
lbl.place()

t = threading.Thread(target=updateMousePos)
t.daemon = True
t.start()
###############################################.

#실행
runas_btn = Button(win)
runas_btn.config(text="실행")
runas_btn.config(width = 6, height=1)
runas_btn.place(x=170, y=420)
runas_btn.config(command = runas)
runas_btn.place()

#초기화
clear_btn = Button(win)
clear_btn.config(text="초기화")
clear_btn.config(width = 6, height=1)
clear_btn.place(x=236, y=420)
clear_btn.config(command = clear)
clear_btn.place()

btn_info = Button(win)
btn_info.config(text="도움")
#btn_info.config(width = 2, height=1)
btn_info.place(x=30, y=420)
btn_info.config(command = infoBox)
btn_info.place()



win.mainloop()
