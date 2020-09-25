from ppadb.client import Client as AdbClient
from bs4 import BeautifulSoup
import re
from tkinter import*
import pyautogui
import pyperclip
from tkinter import filedialog
from openpyxl import load_workbook
import os
import tkinter.messagebox
import time
import threading

time.sleep(0.07)
pyautogui.hotkey('win', 'r')
time.sleep(0.07)
pyautogui.write('cmd') 
time.sleep(0.07)
pyautogui.press('enter')
time.sleep(1.2)
pyautogui.write('adb shell exit') 
time.sleep(0.07)
pyautogui.press('enter')
time.sleep(0.7)# 엔터를 너무 빨리치면 exit를 너무 빨리 입력해서 안 닫아짐
pyautogui.write('exit') 
time.sleep(0.07)
pyautogui.press('enter')

###############################################
win = Tk()
win.geometry("320x460")
win.title("메엠 템세팅 폰크로 v.1")
win.option_add("*Font","맑은고딕 10")
###############################################


global filepath
global file_Sheet


client = AdbClient(host="127.0.0.1", port=5037)
device = client.devices()[0]



def infoBox():
    asdfoijn = '''[폰크로 사용방법]

■ AddItem 치트 화면에서 3가지 좌표값을 확인해주세요.
 ① Type Here... : 아이템 index를 입력하는 박스 위치
 ② Backspace : ① 선택 후 키패드 출력 시 지우기 버튼 위치
 ③ Execute : 우측 하단 Execute 버튼 위치

□ 좌표값을 확인하는 방법
 1. 개발자옵션 활성화
 2. 검색: 포인터 위치
 3. 활성화 후 X: / Y: 값 확인

■ 좌표값 세팅
 1. 파일찾기: 지정된 양식의 파일을 입력
                 > 좌표값이 C ~ H 열에 6개가 있어야 합니다.
 2. 시트명: 좌표값이 있는 시트명 입력
 3. 행: 문서에서 사용할 디바이스의 행 번호 입력
 4. 좌표값 버튼: 1,2,3번 완료 후 선택 시 자동 입력

■ INDEX 세팅
 1. 파일찾기: 사용할 인덱스 파일 첨부
 2. 시트명: 인덱스가 있는 시트명 입력
 3. 범위지정
  - 행 범위: 생성할 인덱스의 행 범위를 입력(ex. 21~50)
               /단, 반드시 2 이상으로 입력
  - 열: 인덱스가 있는 열 번호 입력 (A = 1 / B = 2 / C = 3...)
 

[주의사항]
 - 상단 Device: 에 아무런 정보도 뜨지 않는다면 연결되지 않은 상태입니다.
 - 첨부한 파일의 용량이 큰 경우 파일을 찾지 못하거나 늦어질 수 있습니다.
   가급적 사용할 시트만 분리하시거나 인덱스만 따로 추출하여 새 파일을 만들어주세요.
 - 속도가 PC 매크로에 비해 5배 정도 느립니다. 많은 양의 세팅은 PC 매크로를 사용해주세요.
 - 일시정지 기능은 없지만 케이블을 뽑으면 매크로가 종료됩니다.'''
    tkinter.messagebox.showinfo("설명서", asdfoijn)
###############################################
def Device_openFile():
    global Device_filepath
    Device_filepath = filedialog.askopenfilename()
    Device_fileLabel.config(text = os.path.basename(Device_filepath))
    print(Device_filepath)


def device_location():
    global Device_filepath
    Device_fise_Sheet = Device_sheetEnt.get()

    Device_load_wb = load_workbook(Device_filepath, data_only=True)
    Device_load_ws = Device_load_wb[Device_fise_Sheet]
    phone_device_row = int(device_num_Ent.get())
    #pppp = 3
    #qqqqq = 1
    #device_location_range = range(0,6)
    #for i in device_location_range :
        #location_[qqqqq].insert(load_ws.cell(device_location_range,phone_device_row).value)
        #pppp+=1
        #qqqqq +=1
    
    type_here_xlct.delete(0,1000)
    type_here_ylct.delete(0,1000)
    backSpace_xlct.delete(0,1000)
    backSpace_ylct.delete(0,1000)
    execute_xlct.delete(0,1000)
    execute_ylct.delete(0,1000)
    
    lct_num1 = Device_load_ws.cell(phone_device_row,3).value
    lct_num2 = Device_load_ws.cell(phone_device_row,4).value
    lct_num3 = Device_load_ws.cell(phone_device_row,5).value
    lct_num4 = Device_load_ws.cell(phone_device_row,6).value
    lct_num5 = Device_load_ws.cell(phone_device_row,7).value
    lct_num6 = Device_load_ws.cell(phone_device_row,8).value
    
    type_here_xlct.insert(0, lct_num1)
    type_here_ylct.insert(0, lct_num2)
    backSpace_xlct.insert(0, lct_num3)
    backSpace_ylct.insert(0, lct_num4)
    execute_xlct.insert(0, lct_num5)
    execute_ylct.insert(0, lct_num6)
    
###############################################


###############################################
def update_device():
    global lbl
    global text
    while True:
        text.set(client.devices())
        time.sleep(1)
        
###############################################    
def openFile():
    global filepath
    filepath = filedialog.askopenfilename()
    fileLabel.config(text = os.path.basename(filepath))
    print(filepath)
###############################################
###############################################
def runas():
    TYPE_HERE_X = int(type_here_xlct.get())
    TYPE_HERE_Y = int(type_here_ylct.get())

    BACK_SPACE_X = int(backSpace_xlct.get())
    BACK_SPACE_Y = int(backSpace_ylct.get())

    #COMPLETE_X = int(complete_xlct.get())
    #COMPLETE_Y = int(complete_ylct.get())

    EXECUTE_X = int(execute_xlct.get())
    EXECUTE_Y = int(execute_ylct.get())
    
    global filepath
    fise_Sheet = sheetEnt.get()
    
    load_wb = load_workbook(filepath, data_only=True)
    load_ws = load_wb[fise_Sheet]
    
    index_list = []
    cul = int(culumm_Ent.get()) -1
    row = int(row_Ent.get())
    
    cul2 = int(culumm_Ent2.get())
    cul_range = (cul2 - cul)+1
    
    a = range(0,cul_range)
    
    for i in a:
        index_list.append(load_ws.cell(cul,row).value)
        cul +=1
       
    for j in a:
        j+=1 
        
        device.shell(f'input tap {TYPE_HERE_X} {TYPE_HERE_Y}');
        time.sleep(0.07)
        device.shell(f'input swipe {BACK_SPACE_X} {BACK_SPACE_Y} {BACK_SPACE_X} {BACK_SPACE_Y} 1000')
        time.sleep(0.07)
        device.shell(f'input text {index_list[j]}')  
        time.sleep(0.07)
        #device.shell(f'input tap {COMPLETE_X} {COMPLETE_Y}'); 
        device.shell('input keyevent 66');
        time.sleep(0.07)
        device.shell(f'input tap {EXECUTE_X} {EXECUTE_Y}');
        time.sleep(0.07)
        
        
    
###############################################
# 세팅 라벨
mobileSET_Label = Label(win)
mobileSET_Label.config(text = "[Setting]")
mobileSET_Label.config(font=("맑은고딕", 11))
mobileSET_Label.place(x=20, y=25)
mobileSET_Label.place()

#type here좌표 라벨
type_here_label = Label(win)
type_here_label.config(text = "▶ Type Here...")
type_here_label.place(x=25, y=105)
type_here_label.place()

#type here좌표 라벨
type_here_xlabel = Label(win)
type_here_xlabel.config(text = "X:")
type_here_xlabel.place(x=125, y=105)
type_here_xlabel.place()

#nox x 입력창
type_here_xlct = Entry(win)
type_here_xlct.place(x=145, y=105, width=40, height=18)
type_here_xlct.place()

#nox  y좌표 라벨
type_here_ylabel = Label(win)
type_here_ylabel.config(text = "Y:")
type_here_ylabel.place(x=190, y=105)
type_here_ylabel.place()

#nox y 입력창
type_here_ylct = Entry(win)
type_here_ylct.place(x=210, y=105, width=40, height=18)
type_here_ylct.place()

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@###


#Back space 좌표 라벨
backSpace_label = Label(win)
backSpace_label.config(text = "▶ Backspace")
backSpace_label.place(x=25, y=125)
backSpace_label.place()

#type here좌표 라벨
backSpace_xlabel = Label(win)
backSpace_xlabel.config(text = "X:")
backSpace_xlabel.place(x=125, y=125)
backSpace_xlabel.place()

#nox x 입력창
backSpace_xlct = Entry(win)
backSpace_xlct.place(x=145, y=125, width=40, height=18)
backSpace_xlct.place()

#nox  y좌표 라벨
backSpace_ylabel = Label(win)
backSpace_ylabel.config(text = "Y:")
backSpace_ylabel.place(x=190, y=125)
backSpace_ylabel.place()

#nox y 입력창
backSpace_ylct = Entry(win)
backSpace_ylct.place(x=210, y=125, width=40, height=18)
backSpace_ylct.place()

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@###
#Back space 좌표 라벨
#complete_label = Label(win)
#complete_label.config(text = "▶ 완료")
#complete_label.place(x=25, y=95)
#complete_label.place()

#type here좌표 라벨
#complete_xlabel = Label(win)
#complete_xlabel.config(text = "X:")
#complete_xlabel.place(x=125, y=95)
#complete_xlabel.place()

#nox x 입력창
#complete_xlct = Entry(win)
#complete_xlct.place(x=145, y=95, width=40, height=18)
#complete_xlct.place()

#nox  y좌표 라벨
#complete_ylabel = Label(win)
#complete_ylabel.config(text = "Y:")
#complete_ylabel.place(x=190, y=95)
#complete_ylabel.place()

#nox y 입력창
#complete_ylct = Entry(win)
#complete_ylct.place(x=210, y=115, width=40, height=18)
#complete_ylct.place()



#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@###
#Execute좌표 라벨
execute_label = Label(win)
execute_label.config(text = "▶ Execute")
execute_label.place(x=25, y=145)
execute_label.place()

#type here좌표 라벨
execute_xlabel = Label(win)
execute_xlabel.config(text = "X:")
execute_xlabel.place(x=125, y=145)
execute_xlabel.place()

#nox x 입력창
execute_xlct = Entry(win)
execute_xlct.place(x=145, y=145, width=40, height=18)
execute_xlct.place()

#nox  y좌표 라벨
execute_ylabel = Label(win)
execute_ylabel.config(text = "Y:")
execute_ylabel.place(x=190, y=145)
execute_ylabel.place()

#nox y 입력창
execute_ylct = Entry(win)
execute_ylct.place(x=210, y=145, width=40, height=18)
execute_ylct.place()

###############################################
#파일 찾는 버튼        
fileLoad_btn = Button(text = "파일찾기", command = openFile)
fileLoad_btn.place(x=25, y=181)
fileLoad_btn.place()

#불러온 파일 이름 라벨
fileLabel2 = Label(win)
fileLabel2.config(text = "파일:")
fileLabel2.place(x=25, y=206)
fileLabel2.place()
#불러온 파일 이름 라벨2
fileLabel = Label(win)
fileLabel.config(text = "↑ INDEX 파일첨부")
fileLabel.place(x=60, y=206)
fileLabel.place()

###############################################
#시트 라벨
sheetLabel = Label(win)
sheetLabel.config(text = "시트명:")
sheetLabel.place(x=97, y=184)
sheetLabel.place()

#시트 입력창
sheetEnt = Entry(win)
sheetEnt.insert(0,"Sheet1")
sheetEnt.place(x=148, y=186, width=90, height=15)
sheetEnt.place()
###############################################.
#범위 지정 라벨
range_label = Label(win)
range_label.config(text = "범위 지정(※행은 최소 2이상 입력)")
range_label.place(x=24, y=230)
range_label.place()

#행 라벨
culumm_label = Label(win)
culumm_label.config(text = " 행  :")
culumm_label.place(x=26, y=250)
culumm_label.place()

#행 입력창
culumm_Ent = Entry(win)
culumm_Ent.place(x=63, y=250, width=35, height=18)
culumm_Ent.place()

#행2 라벨
culumm_label2 = Label(win)
culumm_label2.config(text = "~")
culumm_label2.place(x=101, y=250)
culumm_label2.place()

#행2 입력창
culumm_Ent2 = Entry(win)
culumm_Ent2.place(x=120, y=250, width=35, height=18)
culumm_Ent2.place()

#열 라벨
row_label = Label(win)
row_label.config(text = " 열  :")
row_label.place(x=26, y=270)
row_label.place()

#열 입력창
row_Ent = Entry(win)
row_Ent.place(x=63, y=270, width=35, height=18)
row_Ent.place()

###############################################.
device_info = Label(win)
device_info.config(text = "Device:")
device_info.place(x=5, y=0)
device_info.place()
###############################################.
text = StringVar()
lbl = Label(win, textvariable=text)
lbl.place(x=55, y=0)
lbl.place()

t = threading.Thread(target=update_device)
t.daemon = True
t.start()
###############################################.
btn_info = Button(win)
btn_info.config(text="도움")
#btn_info.config(width = 2, height=1)
btn_info.place(x=29, y=420)
btn_info.config(command = infoBox)
btn_info.place()

#실행
runas_btn = Button(win)
runas_btn.config(text="실행")
runas_btn.config(width = 6, height=1)
runas_btn.place(x=236, y=420)
runas_btn.config(command = runas)
runas_btn.place()



#메모장 라벨
mouse_location = Label(win)
mouse_location.config(text = "메모장")
mouse_location.place(x=24, y=306)
mouse_location.place()
###############################################################

#파일 찾는 버튼        
Device_fileLoad_btn = Button(text = "파일찾기", command = Device_openFile)
Device_fileLoad_btn.place(x=25, y=48)
Device_fileLoad_btn.place()

#불러온 파일 이름 라벨
Device_fileLabel2 = Label(win)
Device_fileLabel2.config(text = "파일:")
Device_fileLabel2.place(x=25, y=74)
Device_fileLabel2.place()
#불러온 파일 이름 라벨2
Device_fileLabel = Label(win)
Device_fileLabel.config(text = "↑ 좌표파일 첨부")
Device_fileLabel.place(x=60, y=74)
Device_fileLabel.place()

#시트 라벨
Device_sheetLabel = Label(win)
Device_sheetLabel.config(text = "시트명:")
Device_sheetLabel.place(x=97, y=50)
Device_sheetLabel.place()

#시트 입력창
Device_sheetEnt = Entry(win)
Device_sheetEnt.insert(0,"Sheet1")
Device_sheetEnt.place(x=148, y=51, width=50, height=15)
Device_sheetEnt.place()



#행 라벨
sheetLabel = Label(win)
sheetLabel.config(text = "행:")
sheetLabel.place(x=200, y=50)
sheetLabel.place()
#행 번호 입력칸d
device_num_Ent = Entry(win)
device_num_Ent.place(x=225, y=51, width=30, height=15)
device_num_Ent.place()

#실행
device_lct_btn = Button(win)
device_lct_btn.config(text="좌표값")
device_lct_btn.config(width = 6, height=1)
device_lct_btn.place(x=197, y=71)
device_lct_btn.config(command = device_location)
device_lct_btn.place()
################################################################


textBox = Text(win)
textBox.place(x=26, y=325, width=269, height=80)
textBox.place()


win.mainloop()
