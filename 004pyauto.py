#녹스 리저브리스트 매크로

pip install pyautogui  
pip install pyperclip

import time
import pyperclip
import pyautogui 
from openpyxl import load_workbook

load_wb = load_workbook("C:/Users/user/Desktop/파일이름yo.xlsx", data_only=True)
load_ws = load_wb['slang']

reserve_list = []
a = range(0,101)

for i in a:
    i +=1
    reserve_list.append(load_ws.cell(i,2).value)

pyautogui.moveTo(1412, 532, 0.1) #녹스 위치
pyautogui.click(button='left')

for j in a:
    j+=1 
    pyautogui.press('h') #채팅창 입력 부분 녹스의 단축키
    pyperclip.copy(reserve_list[j])
    pyautogui.hotkey("ctrl", "v")
    #pyautogui.write(reserve_list[j])
    time.sleep(1.2) # 어뷰징 방지용 시간차
    pyautogui.press('enter') 
