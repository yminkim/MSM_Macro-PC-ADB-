from ppadb.client import Client as AdbClient
from bs4 import BeautifulSoup
import re
from tkinter import*
import pyautogui
import pyperclip
from tkinter import filedialog
from openpyxl import load_workbook
import os
import tkinter.messagebox
import time
import threading


###############################################
win = Tk()
win.geometry("320x460")
win.title("폰메엠 템세팅 매크로 v.1")
win.option_add("*Font","맑은고딕 10")
###############################################

client = AdbClient(host="127.0.0.1", port=5037)
device = client.devices()[0]


global filepath
global file_Sheet
###############################################
def update_device():
    global lbl
    global text
    while True:
        text.set(client.devices())
        time.sleep(1)
        
###############################################    
def openFile():
    global filepath
    filepath = filedialog.askopenfilename()
    fileLabel.config(text = os.path.basename(filepath))
    print(filepath)
###############################################
###############################################
def runas():
    global filepath
    fise_Sheet = sheetEnt.get()
    
    load_wb = load_workbook(filepath, data_only=True)
    load_ws = load_wb[fise_Sheet]
    
    index_list = []
    cul = int(culumm_Ent.get()) -1
    row = int(row_Ent.get())
    
    cul2 = int(culumm_Ent2.get())
    cul_range = (cul2 - cul)+1
    
    a = range(0,cul_range)
    
    for i in a:
        index_list.append(load_ws.cell(cul,row).value)
        cul +=1
       
    for j in a:
        j+=1 
        device.shell('input tap 1800 350');
        time.sleep(0.07)
        device.shell('input swipe 2440 1221 2440 1221 1000')
        time.sleep(0.07)
        device.shell(f'input text {index_list[j]}')  
        time.sleep(0.07)
        device.shell('input tap 2384 1350');  
        time.sleep(0.07)
        device.shell('input tap 2269 1341'); 
        time.sleep(0.07)

    
###############################################




###############################################
#파일 찾는 버튼        
fileLoad_btn = Button(text = "파일찾기", command = openFile)
fileLoad_btn.place(x=25, y=160)
fileLoad_btn.place()

#불러온 파일 이름 라벨
fileLabel2 = Label(win)
fileLabel2.config(text = "파일:")
fileLabel2.place(x=25, y=185)
fileLabel2.place()
#불러온 파일 이름 라벨2
fileLabel = Label(win)
fileLabel.config(text = "↑ 파일첨부 해주세요")
fileLabel.place(x=60, y=185)
fileLabel.place()

###############################################
#시트 라벨
sheetLabel = Label(win)
sheetLabel.config(text = "시트명:")
sheetLabel.place(x=97, y=163)
sheetLabel.place()

#시트 입력창
sheetEnt = Entry(win)
sheetEnt.insert(0,"Sheet1")
sheetEnt.place(x=148, y=165, width=90, height=15)
sheetEnt.place()
###############################################.
#범위 지정 라벨
range_label = Label(win)
range_label.config(text = "범위 지정(※행은 최소 2이상 입력)")
range_label.place(x=24, y=215)
range_label.place()

#행 라벨
culumm_label = Label(win)
culumm_label.config(text = " 행  :")
culumm_label.place(x=26, y=235)
culumm_label.place()

#행 입력창
culumm_Ent = Entry(win)
culumm_Ent.place(x=63, y=235, width=35, height=18)
culumm_Ent.place()

#행2 라벨
culumm_label2 = Label(win)
culumm_label2.config(text = "~")
culumm_label2.place(x=101, y=235)
culumm_label2.place()

#행2 입력창
culumm_Ent2 = Entry(win)
culumm_Ent2.place(x=120, y=235, width=35, height=18)
culumm_Ent2.place()

#열 라벨
row_label = Label(win)
row_label.config(text = " 열  :")
row_label.place(x=26, y=255)
row_label.place()

#열 입력창
row_Ent = Entry(win)
row_Ent.place(x=63, y=255, width=35, height=18)
row_Ent.place()

###############################################.
###############################################.
text = StringVar()
lbl = Label(win, textvariable=text)
lbl.place(x=195, y=10)
lbl.place()

t = threading.Thread(target=update_device)
t.daemon = True
t.start()
###############################################.
#실행
runas_btn = Button(win)
runas_btn.config(text="실행")
runas_btn.config(width = 6, height=1)
runas_btn.place(x=170, y=420)
runas_btn.config(command = runas)
runas_btn.place()


win.mainloop()
